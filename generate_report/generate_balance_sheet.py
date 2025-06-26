import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

def generate_balance_sheet(
    excel_path='journal_entry_processed_data/Combine_data/',
    excel_file='Merge',
    ledger_path='account_configuration/Ledger_config_data/Ledger_Library.xlsx',
    output_path='reporting_files/Balance_Sheet/balance_sheet.xlsx'
):
    # Load the journal entry data
    df_data = pd.read_excel(excel_path + excel_file + '.xlsx', engine='openpyxl')

    # Pivot table for trial balance
    trial_balance = pd.pivot_table(
        df_data, index=['J4 - Ledger_Name', 'J2 - Date'],
        values=['J6 - Dr Amount', 'J7 - Cr Amount'],
        aggfunc='sum').fillna(0)

    trial_balance['Balance'] = trial_balance['J6 - Dr Amount'] - trial_balance['J7 - Cr Amount']

    # Create mapping dictionaries
    mapp_dr = trial_balance.groupby('J4 - Ledger_Name')['J6 - Dr Amount'].sum().to_dict()
    mapp_cr = trial_balance.groupby('J4 - Ledger_Name')['J7 - Cr Amount'].sum().to_dict()
    mapp_bl = trial_balance.groupby('J4 - Ledger_Name')['Balance'].sum().to_dict()

    # Normalize keys to lowercase
    mapp_dr = {k.lower(): v for k, v in mapp_dr.items()}
    mapp_cr = {k.lower(): v for k, v in mapp_cr.items()}
    mapp_bl = {k.lower(): v for k, v in mapp_bl.items()}

    # Load ledger data
    df = pd.read_excel(ledger_path, engine='openpyxl')
    df = df[df['C5'].notnull()]
    df['Total Balance'] = df['C5'].map(mapp_bl)
    df = df.dropna(subset=['Total Balance'])

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active

    data = [
        ['Construction Company', " ", 'Closing Balance'],
        ['Balance Sheet as at', " ", 'Current Period'],
        ['31-03-2024', " ", '05/01/15 to 05/01/16'],
        ['', '', ''],
        ['Asset', 'Notes', ' ']
    ]

    note = 0

    # ASSETS
    asset_df = df[df['C1'] == 'ASSET']
    for section in asset_df['C2'].unique():
        section_data = [[section, "", ""]]
        summary = asset_df[asset_df['C2'] == section].groupby('C4')['Total Balance'].sum()
        section_total = summary.sum()
        for ledger, value in summary.items():
            if float(value) > 20000:
                note += 1
                section_data.append([ledger, note, value])
            else:
                section_data.append([ledger, "", value])
        section_data.append(["Total " + section, "", section_total])
        data.extend(section_data)
        data.append(["", "", ""])

    # LIABILITIES
    data.append(['', '', ''])
    data.append(['Liability', 'Notes', ' '])
    liability_df = df[df['C1'] == 'LIABILITY']
    for section in liability_df['C2'].unique():
        section_data = [[section, "", ""]]
        summary = liability_df[liability_df['C2'] == section].groupby('C4')['Total Balance'].sum()
        section_total = summary.sum()
        for ledger, value in summary.items():
            if float(value) > 20000:
                note += 1
                section_data.append([ledger, note, value])
            else:
                section_data.append([ledger, "", value])
        section_data.append(["Total " + section, "", section_total])
        data.extend(section_data)
        data.append(["", "", ""])

    # CAPITAL
    data.append(["", "", ""])
    data.append(["Capital", "Notes", " "])
    capital_df = df[df['C1'] == 'Capital']
    summary = capital_df.groupby('C4')['Total Balance'].sum()
    capital_total = summary.sum()
    for ledger, value in summary.items():
        if float(value) > 20000:
            note += 1
            data.append([ledger, note, value])
        else:
            data.append([ledger, "", value])
    data.append(["Total Equity", "", capital_total])
    data.append(["", "", ""])

    # Write to worksheet
    border_style = Border(top=Side(style='thin'))
    font_style = Font(color="FF0000")

    for i, row in enumerate(data):
        ws.append(row)
        if 'Total' in str(row[0]):
            ws.cell(row=i + 1, column=3).border = border_style

    # Align cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left')
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal='center')
            elif cell.column == 3:
                cell.alignment = Alignment(horizontal='right')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = (max_length + 1)

    # Save workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Balance sheet saved to '{output_path}'")

