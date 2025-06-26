def generate_detail_balance_sheets():
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Font, Alignment
    import os

    excel_path = 'journal_entry_processed_data/Combine_data/'
    excel_file = 'Merge'
    fl_name = excel_file.split('_')[0]

    df_data = pd.read_excel(excel_path + excel_file + '.xlsx', engine='openpyxl')

    trial_balance = pd.pivot_table(df_data, index=['J4 - Ledger_Name'], 
                                   values=['J6 - Dr Amount', 'J7 - Cr Amount'], 
                                   aggfunc='sum').fillna(0)
    trial_balance['Balance'] = trial_balance['J6 - Dr Amount'] - trial_balance['J7 - Cr Amount']

    print("sum of debit", trial_balance['J6 - Dr Amount'].sum())
    print("sum of credit", trial_balance['J7 - Cr Amount'].sum())

    data2 = pd.pivot_table(df_data, index=['J4 - Ledger_Name', 'J2 - Date'], 
                                   values=['J6 - Dr Amount', 'J7 - Cr Amount'], 
                                   aggfunc='sum').fillna(0)
    data2['Balance'] = (data2['J6 - Dr Amount'] - data2['J7 - Cr Amount'])

    account_summary = data2.groupby(['J4 - Ledger_Name', 'J2 - Date']).agg({
        'J6 - Dr Amount': 'sum', 
        'J7 - Cr Amount': 'sum', 
        'Balance': 'sum'
    })

    print("Account Summary:")
    print(account_summary)

    mapp_bl = trial_balance['Balance'].groupby(trial_balance.index).sum().to_dict()
    mapp_bl = {k.lower(): v for k, v in mapp_bl.items()}

    main_df_path = "account_configuration/Ledger_config_data/Ledger_Library.xlsx"   
    df = pd.read_excel(main_df_path, engine='openpyxl')
    df = df[df['C5'].notnull()]
    df['Total Balance'] = df['C5'].map(mapp_bl)
    df = df.dropna(subset=['Total Balance'])

    def export_by_ledger_type(df, ledger_type, output_name):
        sub_df = df[df['C1'] == ledger_type]
        a = sub_df.groupby(['C4', 'C5'])['Total Balance'].value_counts()
        index = a.keys()
        result_dict = {}
        for idx in index:
            if idx[0] not in result_dict:
                result_dict[idx[0]] = {}
            result_dict[idx[0]][idx[1]] = idx[2]

        wb = Workbook()
        ws = wb.active
        data = []

        for i in result_dict:
            sum_s = 0
            c2_data = [[i, '         ']]
            for j in result_dict[i]:
                c2_data.append([j, result_dict[i][j]])
                sum_s += result_dict[i][j]
            c2_data.append(["Total ", sum_s])
            c2_data.append(["", ""])
            c2_data.append(["", ""])
            data.extend(c2_data)

        border_style1 = Border(top=Side(style='thin'))

        for i, row in enumerate(data):
            ws.append(row)
            if 'C4' in row[0]:
                ws.cell(row=i + 1, column=1).font = Font(bold=True)
            elif 'Total' in row[0]:
                ws.cell(row=i + 1, column=2).border = border_style1

        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='bottom')
                elif cell.column == 2:
                    cell.alignment = Alignment(horizontal='right', vertical='bottom')

        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        balance_path = 'reporting_files/Detail_Balance_Sheet/'
        wb.save(balance_path + fl_name + output_name)

    # Generate all 5 sheets
    export_by_ledger_type(df, 'EXPENSE', "_detail_expense_balance_sheet.xlsx")
    export_by_ledger_type(df, 'ASSET', "_detail_asset_balance_sheet.xlsx")
    export_by_ledger_type(df, 'LIABILITY', "_detail_liability_balance_sheet.xlsx")
    export_by_ledger_type(df, 'REVENUE', "_detail_revenue_balance_sheet.xlsx")
    export_by_ledger_type(df, 'Capital', "_detail_capital_balance_sheet.xlsx")
