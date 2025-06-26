import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

def generate_pnl_report():
    excel_path = 'journal_entry_processed_data/Combine_data/'
    excel_file = 'Merge'
    fl_name = excel_file.split('_')[0]

    df_data = pd.read_excel(excel_path + excel_file + '.xlsx', engine='openpyxl')

    trial_balance = pd.pivot_table(df_data, index=['J4 - Ledger_Name'], 
                                   values=['J6 - Dr Amount', 'J7 - Cr Amount'], 
                                   aggfunc='sum').fillna(0)

    trial_balance['Balance'] = trial_balance['J6 - Dr Amount'] - trial_balance['J7 - Cr Amount']

    print("sum of debit", trial_balance['J6 - Dr Amount'].sum())
    print("sum of credit", trial_balance['J7 - Cr Amount'].sum())

    data2 = pd.pivot_table(df_data, index=['J4 - Ledger_Name','J2 - Date'], 
                           values=['J6 - Dr Amount', 'J7 - Cr Amount'], 
                           aggfunc='sum').fillna(0)
    data2['Balance'] = (data2['J6 - Dr Amount'] - data2['J7 - Cr Amount'])
    account_summary = data2.groupby(['J4 - Ledger_Name','J2 - Date']).agg({'J6 - Dr Amount': 'sum', 'J7 - Cr Amount': 'sum', 'Balance': 'sum'})
    print("Account Summary:")
    print(account_summary)

    dr = trial_balance.groupby('J4 - Ledger_Name')['J6 - Dr Amount'].sum()
    ledger = list(dr.keys())
    amount = list(dr.values)
    mapp_dr = {ledger[i].lower(): amount[i] for i in range(len(ledger))}

    cr = trial_balance.groupby('J4 - Ledger_Name')['J7 - Cr Amount'].sum()
    ledger = list(cr.keys())
    amount = list(cr.values)
    mapp_cr = {ledger[i].lower(): amount[i] for i in range(len(ledger))}

    bl = trial_balance.groupby('J4 - Ledger_Name')['Balance'].sum()
    ledger = list(bl.keys())
    amount = list(bl.values)
    mapp_bl = {ledger[i].lower(): amount[i] for i in range(len(ledger))}

    main_df_path = 'account_configuration/Ledger_config_data/Ledger_Library.xlsx'   
    df = pd.read_excel(main_df_path, engine='openpyxl')
    df = df[df['C5'].notnull()]

    df['Total Balance'] = df['C5'].map(mapp_bl)
    df = df.dropna(subset=['Total Balance'])

    expense_df = df[df['C1'] == 'EXPENSE']
    expense_total = df[df['C1'] == 'EXPENSE']['Total Balance'].sum()
    as_list = list(expense_df['C2'].unique())

    note = 0
    data = []
    org_data = ['Construction Company'," ", 'Closing Balance']
    bs_sheet = ['PNL as at '," ", 'Current Period']
    Date  = ['31-03-2024'," ",'31-03-2024']
    st =  ["Expense", "Notes ", " "]

    data.append(org_data)
    data.append(bs_sheet)
    data.append(Date)
    data.append(["","",""])
    data.append(st)

    for i in range(len(as_list)):
        c2_data = []
        c2_data.append([as_list[i], ""," "])

        c2_1 = expense_df[expense_df['C2'] == as_list[i]]
        summary = c2_1.groupby('C4')['Total Balance'].sum()
        sum_value = expense_df[expense_df['C2'] == as_list[i]]['Total Balance'].sum()

        for index, value in summary.items():
            if float(value) > 20000:
                note += 1
                c2_data.append([index, note, value])
            else:
                c2_data.append([index, " ", value])

        c2_data.append(["Total " + as_list[i], " ", sum_value])
        data.extend(c2_data)
        data.append(["","",""])

    revenue_df = df[df['C1'] == 'REVENUE']
    revenue_total = df[df['C1'] == 'REVENUE']['Total Balance'].sum()
    as_list = list(revenue_df['C2'].unique())
    data.append(["","",""])
    st2 = ["Revenue", "Notes ", " "]
    data.append(st2)

    for i in range(len(as_list)):
        c2_data = []
        c2_data.append([as_list[i], ""," "])

        c2_1 = revenue_df[revenue_df['C2'] == as_list[i]]
        summary = c2_1.groupby('C4')['Total Balance'].sum()
        sum_value = revenue_df[revenue_df['C2'] == as_list[i]]['Total Balance'].sum()

        for index, value in summary.items():
            if float(value) > 20000:
                note += 1
                c2_data.append([index, note, value])
            else:
                c2_data.append([index, " ", value])

        c2_data.append(["Total " + as_list[i], " ", sum_value])
        data.extend(c2_data)
        data.append(["","",""])

    data.append(["","",""])
    data.append(["","",""])
    data.append(["Total Expense", "", expense_total])
    data.append(["Total Revenue", "", revenue_total])
    data.append(["Net Income", "", expense_total - revenue_total])

    wb = Workbook()
    ws = wb.active

    border_style = Border(top=Side(style='thin'))
    font_style = Font(color="FF0000")

    for i, row in enumerate(data):
        ws.append(row)
        if ('Total' in row[0] or 'Net' in row[0]) and ('Total Expense' not in row[0] and 'Total Revenue' not in row[0]):
            print(row)
            ws.cell(row=i + 1, column=3).border = border_style

    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='bottom')
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.column == 3:
                cell.alignment = Alignment(horizontal='right', vertical='bottom')

    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length) * 1
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    pnl_path = 'reporting_files/PNL/'
    wb.save(pnl_path + fl_name + "PNL.xlsx")
